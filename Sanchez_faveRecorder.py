import openpyxl as op
import os

workbook = op.Workbook()
sheet = workbook.active

sheet["A1"] = "ID"
sheet["B1"] = "First Name"
sheet["C1"] = "Last Name"
sheet["D1"] = "Birth Year"
sheet["E1"] = "Age"
sheet["A2"] = 1
sheet["A3"] = 2
sheet["A4"] = 3

workbook.save("favorite_people.xlsx")

print("\nFavorite Person 1")
f_name1 = input("Enter first name: ")
l_name1 = input("Enter last name: ")
birth1 = int(input("Enter birth year: "))

print("\nFavorite Person 2")
f_name2 = input("Enter first name: ")
l_name2 = input("Enter last name: ")
birth2 = int(input("Enter birth year: "))

print("\nFavorite Person 3")
f_name3 = input("Enter first name: ")
l_name3 = input("Enter last name: ")
birth3 = int(input("Enter birth year: "))

print("\nFavorite people recorded successfully!")

age1 = 2026 - birth1
age2 = 2026 - birth2
age3 = 2026 - birth3

wbk = op.load_workbook("favorite_people.xlsx")
sheet = wbk.active

print("\n=== FAVORITE PEOPLE ===\n")

sheet["B2"] = f_name1
sheet["C2"] = l_name1
sheet["D2"] = birth1
sheet["E2"] = age1
sheet["B3"] = f_name2
sheet["C3"] = l_name2
sheet["D3"] = birth2
sheet["E3"] = age2
sheet["B4"] = f_name3
sheet["C4"] = l_name3
sheet["D4"] = birth3
sheet["E4"] = age3

wbk.save("favorite_people.xlsx")

for rows in sheet.iter_rows(values_only=True):
    print(rows)

os.system("\nPause")