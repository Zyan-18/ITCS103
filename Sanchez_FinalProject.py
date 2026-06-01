import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op


def input_validation():
    pet_name = pname_entry.get()
    age = age_entry.get()
    gender = gender_entry.get()
    pet_type = ptype_entry.get()
    breed = breed_entry.get()
    owner = owner_entry.get()

    if not pet_name or not age or not gender or not pet_type or not breed or not owner:
        messagebox.showerror("Error", "Please fill in all fields.")
        return False
    
    if not age.isdigit():
        messagebox.showerror("Error", "Age must be a number.")
        return False
    
    for answer, name in [(pet_name, "Pet Name"), (gender, "Gender"), (pet_type, "Pet Type"), (breed, "Breed"), (owner, "Owner")]:
        if answer.isdigit():
            messagebox.showerror("Error", f"{name} must be a text, not numbers.")
            return False

    return True


def clear_entries():
    pname_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    gender_entry.set("")
    ptype_entry.delete(0, tk.END)
    breed_entry.delete(0, tk.END)
    owner_entry.delete(0, tk.END)


def display_():
    workbook = op.load_workbook("Sanchez_Database.xlsx")
    sheet = workbook.active

    for item in table.get_children():
        table.delete(item)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)


def append():
    if not input_validation():
        return
    
    pet_name = pname_entry.get()
    age = int(age_entry.get())
    gender = gender_entry.get()
    pet_type = ptype_entry.get()
    breed = breed_entry.get()
    owner = owner_entry.get()
    
    workbook = op.load_workbook("Sanchez_Database.xlsx")
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id, pet_name, age, gender, pet_type, breed, owner])

    messagebox.showinfo("Success", "New pet information added successfully!")

    workbook.save("Sanchez_Database.xlsx")
    display_()
    clear_entries()


def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, "values")

    if values:
        pname_entry.delete(0, tk.END)
        age_entry.delete(0, tk.END)
        gender_entry.delete(0, tk.END)
        ptype_entry.delete(0, tk.END)
        breed_entry.delete(0, tk.END)
        owner_entry.delete(0, tk.END)

        pname_entry.insert(0, values[1])
        age_entry.insert(0, values[2])
        gender_entry.insert(0, values[3])
        ptype_entry.insert(0, values[4])
        breed_entry.insert(0, values[5])
        owner_entry.insert(0, values[6])


def update_():
    selected = table.focus()
    
    if not input_validation():
        return

    if not selected:
        messagebox.showerror("Error", "Please select your pet information to update.")
        return

    values = table.item(selected, "values")
    id = values[0]

    workbook = op.load_workbook("Sanchez_Database.xlsx")
    sheet = workbook.active

    if values:
    
        for row in sheet.iter_rows(min_row=2):

            if str(row[0].value) == str(id):
                row[1].value = pname_entry.get()
                row[2].value = int(age_entry.get())
                row[3].value = gender_entry.get()
                row[4].value = ptype_entry.get()
                row[5].value = breed_entry.get()
                row[6].value = owner_entry.get()

        workbook.save("Sanchez_Database.xlsx")
        display_()
        clear_entries()

    messagebox.showinfo("Success", "Pet information has been updated successfully!")


def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Please select a record first.")
        return
    
    values = table.item(selected, "values")
    id_rec = values[0]

    confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this record?")
    if not confirm:
        return

    workbook = op.load_workbook("Sanchez_Database.xlsx")
    sheet = workbook.active

    for a, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(id_rec):
            sheet.delete_rows(a)
            break
            
    workbook.save("Sanchez_Database.xlsx")

    messagebox.showinfo("Success", "Record has been deleted successfully!")
    display_()
    clear_entries()

    return


window = tk.Tk()
window.title("Pet Information System")
window.geometry("1425x525")
window.resizable(False, False)
window.configure(bg="#FFF7AF")

title = tk.Label(window, text="Petfolio", font=("Comic Sans MS", 30, "bold"), bg="#FFF7AF", fg="#FF5DB4")
title.grid(row=0, column=0, columnspan=6)

frame = tk.Frame(window, bg="#9EECF5", bd=2, relief="ridge")
frame.grid(row=1, column=0, columnspan=6, padx=15, pady=15)

pname_entry = tk.Entry(frame, font=("Verdana", 13))
pname_entry.grid(row=2, column=1, columnspan=2, padx=12, pady=(12, 1))

pname_label = tk.Label(frame, text="Pet Name", font=("Verdana", 10), bg="#9EECF5")
pname_label.grid(row=3, column=1, columnspan=2)

age_entry = tk.Entry(frame, font=("Verdana", 13))
age_entry.grid(row=2, column=3, columnspan=2, padx=12, pady=(12, 1))

age_label = tk.Label(frame, text="Age (months)", font=("Verdana", 10), bg="#9EECF5")
age_label.grid(row=3, column=3, columnspan=2)

gender_entry = ttk.Combobox(frame, font=("Verdana", 13), state="readonly")
gender_entry['values'] = ("Male", "Female", "Other")
gender_entry.grid(row=2, column=5, columnspan=2, padx=12, pady=(12, 1))

gender_label = tk.Label(frame, text="Gender", font=("Verdana", 10), bg="#9EECF5")
gender_label.grid(row=3, column=5, columnspan=2)

ptype_entry = tk.Entry(frame, font=("Verdana", 13))
ptype_entry.grid(row=4, column=1, columnspan=2, padx=12, pady=(12, 1))

ptype_label = tk.Label(frame, text="Pet Type", font=("Verdana", 10), bg="#9EECF5")
ptype_label.grid(row=5, column=1, columnspan=2)

breed_entry = tk.Entry(frame, font=("Verdana", 13))
breed_entry.grid(row=4, column=3, columnspan=2, padx=12, pady=(12, 1))

breed_label = tk.Label(frame, text="Breed", font=("Verdana", 10), bg="#9EECF5")
breed_label.grid(row=5, column=3, columnspan=2)

owner_entry = tk.Entry(frame, font=("Verdana", 13))
owner_entry.grid(row=4, column=5, columnspan=2, padx=12, pady=(12, 1))

owner_label = tk.Label(frame, text="Owner", font=("Verdana", 10), bg="#9EECF5")
owner_label.grid(row=5, column=5, columnspan=2)


submit_btn = tk.Button(window, text="Submit", font=("Verdana", 12, "bold"), bg="pink1", command=append)
submit_btn.grid(row=6, column=1, padx=(90, 0), pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Verdana", 12, "bold"), bg="yellow", command=update_)
update_btn.grid(row=6, column=2, padx=(140, 0), pady=(10, 20))

delete_btn = tk.Button(window, text="Delete", bg="lightblue", fg="black",font=("Verdana", 12, "bold"), command=delete)
delete_btn.grid(row=6, column=3, padx=(140, 0), pady=(10, 20))


table = ttk.Treeview(window,
    columns=("Pet ID", "Pet Name", "Age", "Gender", "Pet Type", "Breed", "Owner"),
    show="headings")

for headings in ("Pet ID", "Pet Name", "Age", "Gender", "Pet Type", "Breed", "Owner"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

display_()

window.mainloop()