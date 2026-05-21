import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op

def display_content():
    wbk = op.load_workbook("ordersDB.xlsx")
    sheet = wbk.active

    for item in table.get_children():
            table.delete(item)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)

def input_validation():
    name = cname_entry.get()
    product = product_entry.get()
    quantity = qty_entry.get()
    price = price_entry.get()

    if not name or not product or not quantity or not price:
        messagebox.showerror("Error", "Please fill all fields.")
        return False

    if not quantity.isdigit() or not price.isdigit():
        messagebox.showerror("Error", "Quantity and Price must be numbers.")
        return False
    
    return True

def append_():
     if not input_validation():
        return
     
     cname = cname_entry.get()
     product = product_entry.get()
     qty = qty_entry.get()
     price = price_entry.get()
     total_amount = qty * price
     
     wbk = op.load_workbook("ordersDB.xlsx")
     sheet = wbk.active

     new_ID = sheet.max_row

     sheet.append([new_ID, cname, product, qty, price])

     messagebox.showinfo("Success", "Order added successfully")

     wbk.save("ordersDB.xlsx")
     display_content()


def auto_populate(event):
    selected = table.focus()
    values = table.item(selected,"values")

    if values:
        cname_entry.delete(0,tk.END)
        product_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        price_entry.delete(0,tk.END)

        cname_entry.insert(0,values[1])
        product_entry.insert(0,values[2])
        qty_entry.insert(0,values[3])
        price_entry.insert(0,values[4])

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error","Select a record first!")

    values = table.item(selected,"values")
    record_id = values[0]

    confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete this record?")
    
    if not confirm:
        return

    workbook = op.load_workbook("ordersDB.xlsx")
    sheet = workbook.active

    for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break

    workbook.save("excelDB.xlsx")
    messagebox.showinfo("Success","Record deleted successfully")
    display_content


window = tk.Tk()
window.title("Simple Ordering System")
window.configure(bg="lightblue")

# Form Title
title = tk.Label(window, text="Simple Ordering System", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = tk.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = tk.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="lightblue")
cname_label.grid(row=3, column=1, columnspan=2)

# Product Entry
product_entry = tk.Entry(genframe, font=("Poppins", 12))
product_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

product_label = tk.Label(genframe, text="Product", font=("Poppins", 10, "italic"), bg="lightblue")
product_label.grid(row=3, column=3, columnspan=2)

# Quantity Entry
qty_entry = tk.Entry(genframe, font=("Poppins", 12))
qty_entry.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

qty_label = tk.Label(genframe, text="Quantity", font=("Poppins", 10, "italic"), bg="lightblue")
qty_label.grid(row=5, column=1, columnspan=2)

# Price Entry
price_entry = tk.Entry(genframe, font=("Poppins", 12))
price_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

price_label = tk.Label(genframe, text="Price", font=("Poppins", 10, "italic"), bg="lightblue")
price_label.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink", command=append_)
submit_btn.grid(row=6, column=1, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen")
update_btn.grid(row=6, column=2)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"), command=delete)
delete_btn.grid(row=6, column=3)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Product", "Quantity", "Price", "Total"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

display_content()

window.mainloop()